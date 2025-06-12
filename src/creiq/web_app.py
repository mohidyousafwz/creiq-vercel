"""
Web Dashboard for CREIQ Data Extraction Service.
"""
import sys
import asyncio
import json
import secrets
import csv
import io
import zipfile
import tempfile
from datetime import datetime, timedelta, timezone
from typing import Optional, Dict, Any, List
from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment

# Fix for Windows asyncio issues
if sys.platform == 'win32':
    asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())

from fastapi import FastAPI, Request, Response, Form, Depends, HTTPException, UploadFile, File, BackgroundTasks
from fastapi.responses import HTMLResponse, StreamingResponse, RedirectResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from starlette.middleware.sessions import SessionMiddleware
from sqlalchemy.orm import Session
from sqlalchemy import func, desc
from pydantic import BaseModel
import uvicorn

from src.creiq.config.settings import (
    PASSCODE, SESSION_DURATION_DAYS, SECRET_KEY,
    API_HOST, API_PORT, API_RELOAD, BROWSER_HEADLESS, RESULTS_DIR
)
from src.creiq.database.database import get_db, engine, Base
from src.creiq.database.models import RollNumber, Appeal
from src.creiq.database.service import DatabaseService
from src.creiq.utils.logger import logger
from src.creiq.utils.roll_number_reader import read_roll_numbers_from_csv
from src.creiq.services.extraction_service import ExtractionService
import threading
from src.creiq.playwright_automation import PlaywrightAutomation

# Create FastAPI app
app = FastAPI(title="CREIQ Dashboard", version="1.0.0")

# Add session middleware
app.add_middleware(SessionMiddleware, secret_key=SECRET_KEY)

# Mount static files
static_path = Path(__file__).parent.parent.parent / "static"
static_path.mkdir(exist_ok=True)
app.mount("/static", StaticFiles(directory=str(static_path)), name="static")

# Setup templates
templates_path = Path(__file__).parent.parent.parent / "templates"
templates = Jinja2Templates(directory=str(templates_path))

# Global variables for real-time updates
active_extractions: Dict[str, Dict[str, Any]] = {}
extraction_logs: List[Dict[str, Any]] = []
log_subscribers: List[asyncio.Queue] = []
shutdown_signal = threading.Event()

# Create database tables
Base.metadata.create_all(bind=engine)


# Jinja2 filters
def format_datetime(value):
    """Format datetime for display."""
    if not value:
        return ""
    if isinstance(value, str):
        value = datetime.fromisoformat(value)
    return value.strftime("%Y-%m-%d %H:%M:%S")


def format_date(value):
    """Format date for display."""
    if not value:
        return ""
    if isinstance(value, str):
        value = datetime.fromisoformat(value)
    return value.strftime("%Y-%m-%d")


templates.env.filters['datetime'] = format_datetime
templates.env.filters['date'] = format_date


# Authentication dependency
async def require_auth(request: Request):
    """Check if user is authenticated."""
    if not request.session.get("authenticated"):
        raise HTTPException(status_code=401, detail="Not authenticated")
    return True


# Logging functions
async def add_log(level: str, message: str, details: Dict[str, Any] = None):
    """Add a log entry and notify subscribers."""
    log_entry = {
        "timestamp": datetime.now().isoformat(),
        "level": level,
        "message": message,
        "details": details or {}
    }
    
    # Keep only last 1000 logs
    extraction_logs.insert(0, log_entry)
    if len(extraction_logs) > 1000:
        extraction_logs.pop()
    
    # Notify all subscribers
    for queue in log_subscribers[:]:
        try:
            await queue.put(log_entry)
        except:
            log_subscribers.remove(queue)


# Routes
@app.get("/", response_class=HTMLResponse)
async def login_page(request: Request):
    """Login page."""
    if request.session.get("authenticated"):
        return RedirectResponse(url="/dashboard", status_code=302)
    return templates.TemplateResponse("login.html", {"request": request})


@app.post("/login")
async def login(request: Request, passcode: str = Form(...)):
    """Handle login."""
    if passcode == PASSCODE:
        request.session["authenticated"] = True
        request.session["login_time"] = datetime.now().isoformat()
        await add_log("INFO", "User logged in successfully")
        return RedirectResponse(url="/dashboard", status_code=302)
    
    await add_log("WARNING", "Failed login attempt")
    return templates.TemplateResponse("login.html", {
        "request": request,
        "error": "Invalid passcode"
    })


@app.get("/logout")
async def logout(request: Request):
    """Handle logout."""
    request.session.clear()
    await add_log("INFO", "User logged out")
    return RedirectResponse(url="/", status_code=302)


@app.get("/dashboard", response_class=HTMLResponse)
async def dashboard(request: Request, db: Session = Depends(get_db), _: bool = Depends(require_auth)):
    """Main dashboard page."""
    # Get statistics
    total_roll_numbers = db.query(func.count(RollNumber.roll_number)).scalar()
    total_appeals = db.query(func.count(Appeal.id)).scalar()
    
    # Get status counts
    status_counts = db.query(
        RollNumber.extraction_status,
        func.count(RollNumber.roll_number)
    ).group_by(RollNumber.extraction_status).all()
    
    # Get recently updated roll numbers (last 7 days)
    seven_days_ago = datetime.now() - timedelta(days=7)
    recent_roll_numbers = db.query(RollNumber).filter(
        RollNumber.updated_at >= seven_days_ago
    ).order_by(desc(RollNumber.updated_at)).limit(10).all()
    
    # Get currently processing
    processing_count = len([e for e in active_extractions.values() if e.get("status") == "processing"])
    
    return templates.TemplateResponse("dashboard.html", {
        "request": request,
        "total_roll_numbers": total_roll_numbers,
        "total_appeals": total_appeals,
        "status_counts": dict(status_counts),
        "recent_roll_numbers": recent_roll_numbers,
        "processing_count": processing_count,
        "active_extractions": active_extractions
    })


@app.get("/roll-numbers", response_class=HTMLResponse)
async def roll_numbers_page(request: Request, db: Session = Depends(get_db), _: bool = Depends(require_auth)):
    """Roll numbers management page."""
    return templates.TemplateResponse("roll_numbers.html", {"request": request})


@app.get("/roll-numbers/{roll_number}", response_class=HTMLResponse)
async def roll_number_detail(
    roll_number: str, 
    request: Request, 
    db: Session = Depends(get_db), 
    _: bool = Depends(require_auth)
):
    """Roll number detail page."""
    try:
        db_service = DatabaseService(db)
        roll_record = db_service.get_roll_number(roll_number)
        
        if not roll_record:
            raise HTTPException(status_code=404, detail="Roll number not found")
        
        return templates.TemplateResponse("roll_number_detail.html", {
            "request": request,
            "roll_number": roll_record
        })
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching roll number details: {e}")
        raise HTTPException(status_code=500, detail="Internal server error")


@app.get("/logs", response_class=HTMLResponse)
async def logs_page(request: Request, _: bool = Depends(require_auth)):
    """Scraper logs page."""
    return templates.TemplateResponse("logs.html", {
        "request": request,
        "initial_logs": extraction_logs[:100]  # Last 100 logs
    })


@app.get("/settings", response_class=HTMLResponse)
async def settings_page(request: Request, _: bool = Depends(require_auth)):
    """Settings page."""
    return templates.TemplateResponse("settings.html", {"request": request})


@app.get("/guide", response_class=HTMLResponse)
async def guide_page(request: Request, _: bool = Depends(require_auth)):
    """User guide page."""
    return templates.TemplateResponse("guide.html", {"request": request})


@app.get("/health-status", response_class=HTMLResponse)
async def health_page(request: Request, db: Session = Depends(get_db), _: bool = Depends(require_auth)):
    """System health page."""
    # Get database status
    try:
        db.execute("SELECT 1")
        db_status = "healthy"
    except:
        db_status = "unhealthy"
    
    # Get extraction service status
    extraction_status = "healthy" if not shutdown_signal.is_set() else "shutdown"
    
    return templates.TemplateResponse("health.html", {
        "request": request,
        "db_status": db_status,
        "extraction_status": extraction_status,
        "active_tasks": len(active_extractions),
        "log_count": len(extraction_logs)
    })


# API endpoints for AJAX calls
@app.get("/api/dashboard/stats")
async def get_dashboard_stats(db: Session = Depends(get_db), _: bool = Depends(require_auth)):
    """Get dashboard statistics."""
    total_roll_numbers = db.query(func.count(RollNumber.roll_number)).scalar()
    total_appeals = db.query(func.count(Appeal.id)).scalar()
    
    # Status breakdown
    status_counts = db.query(
        RollNumber.extraction_status,
        func.count(RollNumber.roll_number)
    ).group_by(RollNumber.extraction_status).all()
    
    # Success rate calculation
    completed = next((count for status, count in status_counts if status == "completed"), 0)
    failed = next((count for status, count in status_counts if status == "failed"), 0)
    total_processed = completed + failed
    success_rate = (completed / total_processed * 100) if total_processed > 0 else 0
    
    # Appeals by status
    appeal_status = db.query(
        Appeal.status,
        func.count(Appeal.id)
    ).group_by(Appeal.status).all()
    
    return {
        "total_roll_numbers": total_roll_numbers,
        "total_appeals": total_appeals,
        "processing_count": len([e for e in active_extractions.values() if e.get("status") == "processing"]),
        "status_breakdown": dict(status_counts),
        "success_rate": round(success_rate, 1),
        "appeal_status": dict(appeal_status)
    }


@app.get("/api/roll-numbers/search")
async def search_roll_numbers(
    q: str = "",
    limit: int = 50,
    offset: int = 0,
    db: Session = Depends(get_db),
    _: bool = Depends(require_auth)
):
    """Search roll numbers."""
    query = db.query(RollNumber)
    
    if q:
        query = query.filter(RollNumber.roll_number.contains(q))
    
    total = query.count()
    roll_numbers = query.limit(limit).offset(offset).all()
    
    return {
        "total": total,
        "roll_numbers": [
            {
                "roll_number": r.roll_number,
                "property_description": r.property_description,
                "municipality": r.municipality,
                "extraction_status": r.extraction_status,
                "appeals_count": len(r.appeals),
                "last_extracted_at": r.last_extracted_at.isoformat() if r.last_extracted_at else None,
                "progress": active_extractions.get(r.roll_number, {}).get("progress", "")
            }
            for r in roll_numbers
        ]
    }


@app.post("/api/roll-numbers/upload")
async def upload_roll_numbers(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _: bool = Depends(require_auth)
):
    """Upload CSV file with roll numbers and add them to database immediately."""
    if not file.filename.endswith('.csv'):
        raise HTTPException(status_code=400, detail="Only CSV files are allowed")
    
    try:
        roll_numbers = await read_roll_numbers_from_csv(file)
        
        if not roll_numbers:
            raise HTTPException(status_code=400, detail="No valid roll numbers found")
        
        # Check for duplicates and get detailed info
        existing_query = db.query(RollNumber).filter(
            RollNumber.roll_number.in_(roll_numbers)
        ).all()
        
        existing_info = {}
        for roll in existing_query:
            existing_info[roll.roll_number] = {
                "status": roll.extraction_status,
                "appeals_count": len(roll.appeals),
                "last_extracted": roll.last_extracted_at.isoformat() if roll.last_extracted_at else None
            }
        
        existing_numbers = set(existing_info.keys())
        new_numbers = [r for r in roll_numbers if r not in existing_numbers]
        
        # Add new roll numbers to database immediately
        db_service = DatabaseService(db)
        for roll_number in new_numbers:
            db_service.create_or_update_roll_number(roll_number)
        
        await add_log("INFO", f"Added {len(new_numbers)} new roll numbers to database: {len(roll_numbers)} total ({len(new_numbers)} new, {len(existing_numbers)} existing)")
        
        return {
            "total": len(roll_numbers),
            "new": len(new_numbers),
            "existing": len(existing_numbers),
            "roll_numbers": roll_numbers,
            "new_roll_numbers": new_numbers,
            "existing_info": existing_info,
            "message": f"Added {len(new_numbers)} new roll numbers to database. Found {len(existing_numbers)} existing ones"
        }
        
    except Exception as e:
        await add_log("ERROR", f"Failed to upload CSV: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# Pydantic models
class ProcessRollNumbersRequest(BaseModel):
    roll_numbers: List[str]


@app.post("/api/roll-numbers/process")
async def process_roll_numbers(
    background_tasks: BackgroundTasks,
    request: ProcessRollNumbersRequest,
    db: Session = Depends(get_db),
    _: bool = Depends(require_auth)
):
    """Start processing roll numbers."""
    # Create task
    import uuid
    task_id = str(uuid.uuid4())
    
    # First, ensure all roll numbers exist in database (add new ones if needed)
    db_service = DatabaseService(db)
    for roll_number in request.roll_numbers:
        existing_roll = db_service.get_roll_number(roll_number)
        if not existing_roll:
            # Create new roll number record
            db_service.create_or_update_roll_number(roll_number)
    
    # Initialize extraction tracking
    for roll_number in request.roll_numbers:
        active_extractions[roll_number] = {
            "task_id": task_id,
            "status": "queued",
            "progress": "Waiting to start...",
            "started_at": datetime.now().isoformat()
        }
    
    # Start background task
    background_tasks.add_task(
        run_extraction_task,
        task_id,
        request.roll_numbers,
        db
    )
    
    await add_log("INFO", f"Started processing {len(request.roll_numbers)} roll numbers")
    
    return {
        "task_id": task_id,
        "roll_numbers": request.roll_numbers
    }


async def run_extraction_task(task_id: str, roll_numbers: List[str], db: Session):
    """Run extraction in background."""
    from concurrent.futures import ThreadPoolExecutor
    import asyncio
    
    def run_sync_extraction():
        """Run the synchronous extraction in a separate thread."""
        automation = None
        try:
            # Initialize automation
            automation = PlaywrightAutomation(
                headless=BROWSER_HEADLESS,
                shutdown_signal=shutdown_signal
            )
            
            # Update status for all roll numbers
            for roll_number in roll_numbers:
                active_extractions[roll_number]["status"] = "processing"
                active_extractions[roll_number]["progress"] = "Initializing browser..."
            
            # Start browser
            active_extractions[roll_numbers[0]]["progress"] = "Starting browser..."
            automation.start_browser()
            
            # Navigate to site
            active_extractions[roll_numbers[0]]["progress"] = "Navigating to ARB website..."
            automation.navigate_to_site()
            
            # Process each roll number
            for i, roll_number in enumerate(roll_numbers):
                if shutdown_signal.is_set():
                    raise Exception("Shutdown requested")
                    
                try:
                    # Update progress
                    active_extractions[roll_number]["status"] = "processing"
                    active_extractions[roll_number]["progress"] = f"Processing {i+1}/{len(roll_numbers)}: Checking existing data..."
                    
                    # Check existing appeals in database
                    db_service = DatabaseService(db)
                    existing_roll = db_service.get_roll_number(roll_number)
                    existing_appeal_numbers = set()
                    
                    # Get extraction progress
                    progress_info = db_service.get_extraction_progress(roll_number)
                    extracted_appeal_numbers = progress_info["extracted_appeal_numbers"]
                    
                    if existing_roll and existing_roll.appeals:
                        existing_appeal_numbers = {appeal.appeal_number for appeal in existing_roll.appeals}
                        active_extractions[roll_number]["progress"] = f"Found {len(existing_appeal_numbers)} existing appeals, {len(extracted_appeal_numbers)} fully extracted"
                    
                    # Update database status
                    db_service.update_roll_number_status(roll_number, "processing")
                    
                    # Enter roll number and search
                    active_extractions[roll_number]["progress"] = "Entering roll number..."
                    automation.enter_roll_number(roll_number)
                    
                    active_extractions[roll_number]["progress"] = "Submitting search..."
                    if not automation.submit_search():
                        raise Exception("Failed to submit search")
                    
                    # Extract current appeal data from the page
                    active_extractions[roll_number]["progress"] = "Extracting appeal information..."
                    current_data = automation.extract_data_to_json(roll_number)
                    
                    # Get current appeal numbers from the page
                    current_appeal_numbers = {appeal["appealnumber"] for appeal in current_data.get("appeal_info", [])}
                    
                    # Update total appeals found
                    db_service.update_extraction_progress(roll_number, total_appeals=len(current_appeal_numbers))
                    
                    # Determine which appeals need to be extracted
                    appeals_to_extract = []
                    for appeal in current_data.get("appeal_info", []):
                        appeal_number = appeal.get("appealnumber")
                        if appeal_number and appeal_number not in extracted_appeal_numbers:
                            appeals_to_extract.append(appeal)
                    
                    if not appeals_to_extract:
                        # All appeals already extracted
                        active_extractions[roll_number]["progress"] = f"All {len(current_appeal_numbers)} appeals already extracted"
                        active_extractions[roll_number]["status"] = "completed"
                        db_service.update_roll_number_status(roll_number, "completed")
                        continue
                    
                    # We have appeals to extract
                    active_extractions[roll_number]["progress"] = f"Need to extract {len(appeals_to_extract)} appeals (out of {len(current_appeal_numbers)} total)"
                    
                    # Process the roll number
                    safe_roll_number_dir = roll_number.replace('/', '_').replace('\\', '_').replace(':', '_')
                    roll_number_results_dir = Path(RESULTS_DIR) / safe_roll_number_dir
                    roll_number_results_dir.mkdir(parents=True, exist_ok=True)
                    
                    # Save the summary data
                    summary_json_path = roll_number_results_dir / "appeal_summary.json"
                    automation.save_json_data(current_data, str(summary_json_path))
                    
                    # Save/update property info
                    property_info = current_data.get("property_info", {})
                    db_service.create_or_update_roll_number(roll_number, property_info)
                    
                    # Extract and save appeals progressively
                    successfully_extracted = 0
                    failed_extractions = []
                    
                    for idx, appeal_summary in enumerate(appeals_to_extract):
                        appeal_number = appeal_summary.get("appealnumber")
                        
                        try:
                            # Update progress
                            active_extractions[roll_number]["progress"] = f"Extracting appeal {idx+1}/{len(appeals_to_extract)}: {appeal_number}"
                            
                            # Extract detailed information for this appeal
                            appeal_detail = automation.extract_single_appeal_detail(appeal_summary, str(roll_number_results_dir))
                            
                            # Save the appeal immediately (progressive saving)
                            db_service.save_single_appeal(roll_number, appeal_summary, appeal_detail)
                            
                            successfully_extracted += 1
                            
                            # Update progress in UI
                            total_extracted = len(extracted_appeal_numbers) + successfully_extracted
                            active_extractions[roll_number]["progress"] = f"Extracted {total_extracted}/{len(current_appeal_numbers)} appeals"
                            
                        except Exception as e:
                            logger.error(f"Failed to extract appeal {appeal_number}: {e}")
                            failed_extractions.append(appeal_number)
                            # Continue with next appeal instead of failing completely
                            continue
                    
                    # Update final status
                    total_extracted = len(extracted_appeal_numbers) + successfully_extracted
                    
                    if failed_extractions:
                        # Partial success
                        active_extractions[roll_number]["status"] = "failed"
                        active_extractions[roll_number]["progress"] = f"Extracted {total_extracted}/{len(current_appeal_numbers)} appeals. Failed: {', '.join(failed_extractions)}"
                        db_service.update_roll_number_status(roll_number, "failed", f"Failed to extract appeals: {', '.join(failed_extractions)}")
                    else:
                        # Complete success
                        active_extractions[roll_number]["status"] = "completed"
                        active_extractions[roll_number]["progress"] = f"Extraction complete - {successfully_extracted} new appeals added ({total_extracted} total)"
                        db_service.update_roll_number_status(roll_number, "completed")
                    
                    # Schedule cleanup of completed extraction after 10 seconds
                    import threading
                    def cleanup_extraction():
                        import time
                        time.sleep(10)
                        if roll_number in active_extractions and active_extractions[roll_number]["status"] == "completed":
                            del active_extractions[roll_number]
                    
                    cleanup_thread = threading.Thread(target=cleanup_extraction)
                    cleanup_thread.daemon = True
                    cleanup_thread.start()
                    
                except Exception as e:
                    # Mark individual roll number as failed
                    active_extractions[roll_number]["status"] = "failed"
                    active_extractions[roll_number]["progress"] = f"Error: {str(e)}"
                    
                    db_service = DatabaseService(db)
                    db_service.update_roll_number_status(roll_number, "failed", str(e))
                    
                    # Schedule cleanup of failed extraction after 10 seconds
                    import threading
                    def cleanup_extraction():
                        import time
                        time.sleep(10)
                        if roll_number in active_extractions and active_extractions[roll_number]["status"] == "failed":
                            del active_extractions[roll_number]
                    
                    cleanup_thread = threading.Thread(target=cleanup_extraction)
                    cleanup_thread.daemon = True
                    cleanup_thread.start()
                    
                    continue
                    
        except Exception as e:
            # Mark all remaining roll numbers as failed
            for roll_number in roll_numbers:
                if roll_number in active_extractions and active_extractions[roll_number]["status"] == "processing":
                    active_extractions[roll_number]["status"] = "failed"
                    active_extractions[roll_number]["progress"] = f"Task error: {str(e)}"
        finally:
            # Clean up browser
            if automation:
                try:
                    automation.close()
                except:
                    pass
    
    try:
        await add_log("INFO", f"Starting extraction task {task_id} for {len(roll_numbers)} roll numbers")
        
        # Run the synchronous extraction in a thread pool
        loop = asyncio.get_event_loop()
        with ThreadPoolExecutor(max_workers=1) as executor:
            await loop.run_in_executor(executor, run_sync_extraction)
            
        await add_log("INFO", f"Extraction task {task_id} completed")
        
    except Exception as e:
        await add_log("ERROR", f"Extraction task {task_id} failed: {str(e)}")


@app.get("/api/roll-numbers/export")
async def export_roll_numbers(
    type: str = "all",  # all or processed
    db: Session = Depends(get_db),
    _: bool = Depends(require_auth)
):
    """Export roll numbers as CSV."""
    query = db.query(RollNumber)
    
    if type == "processed":
        query = query.filter(RollNumber.extraction_status == "completed")
    
    roll_numbers = query.all()
    
    # Create CSV
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Header
    writer.writerow([
        "Roll Number", "Property Description", "Municipality", 
        "Status", "Appeals Count", "Last Extracted"
    ])
    
    # Data
    for r in roll_numbers:
        writer.writerow([
            r.roll_number,
            r.property_description or "",
            r.municipality or "",
            r.extraction_status,
            len(r.appeals),
            r.last_extracted_at.strftime("%Y-%m-%d %H:%M:%S") if r.last_extracted_at else ""
        ])
    
    output.seek(0)
    
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode()),
        media_type="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename=roll_numbers_{type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        }
    )


@app.get("/api/roll-numbers/export-with-appeals")
async def export_roll_numbers_with_appeals(
    type: str = "all",  # all or processed
    db: Session = Depends(get_db),
    _: bool = Depends(require_auth)
):
    """Export roll numbers with appeals as Excel files in a ZIP archive."""
    query = db.query(RollNumber)
    
    if type == "processed":
        query = query.filter(RollNumber.extraction_status == "completed")
    
    roll_numbers = query.all()
    
    # Create a temporary directory for Excel files
    with tempfile.TemporaryDirectory() as temp_dir:
        temp_path = Path(temp_dir)
        
        # Process each roll number
        for roll_record in roll_numbers:
            # Skip roll numbers without appeals
            if not roll_record.appeals:
                continue
                
            # Create Excel workbook
            wb = openpyxl.Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Create Main sheet with roll number info
            main_sheet = wb.create_sheet("Main", 0)
            
            # Style for headers
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Add roll number information to Main sheet
            main_headers = ["Field", "Value"]
            for col, header in enumerate(main_headers, 1):
                cell = main_sheet.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Roll number data
            main_data = [
                ["Roll Number", roll_record.roll_number],
                ["Property Description", roll_record.property_description or ""],
                ["Municipality", roll_record.municipality or ""],
                ["Classification", roll_record.classification or ""],
                ["Neighborhood", roll_record.nbhd or ""],
                ["Status", roll_record.extraction_status],
                ["Total Appeals", len(roll_record.appeals)],
                ["Last Extracted", roll_record.last_extracted_at.strftime("%Y-%m-%d %H:%M:%S") if roll_record.last_extracted_at else ""],
                ["Created", roll_record.created_at.strftime("%Y-%m-%d %H:%M:%S") if roll_record.created_at else ""],
                ["Updated", roll_record.updated_at.strftime("%Y-%m-%d %H:%M:%S") if roll_record.updated_at else ""]
            ]
            
            for row_idx, (field, value) in enumerate(main_data, 2):
                main_sheet.cell(row=row_idx, column=1, value=field)
                main_sheet.cell(row=row_idx, column=2, value=value)
            
            # Adjust column widths for Main sheet
            main_sheet.column_dimensions["A"].width = 25
            main_sheet.column_dimensions["B"].width = 50
            
            # Create a sheet for each appeal
            for appeal_idx, appeal in enumerate(roll_record.appeals, 1):
                # Create sheet name (Excel has 31 char limit for sheet names)
                sheet_name = f"Appeal {appeal_idx}"
                if appeal.appeal_number:
                    # Truncate appeal_number if needed to fit in sheet name
                    appeal_number_short = appeal.appeal_number[:20] if len(appeal.appeal_number) > 20 else appeal.appeal_number
                    sheet_name = f"Appeal {appeal_idx} - {appeal_number_short}"
                
                appeal_sheet = wb.create_sheet(sheet_name)
                
                # Appeal headers
                appeal_headers = ["Field", "Value"]
                for col, header in enumerate(appeal_headers, 1):
                    cell = appeal_sheet.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # Appeal data
                appeal_data = [
                    ["Appeal Number", appeal.appeal_number or ""],
                    ["Appellant", appeal.appellant or ""],
                    ["Representative", appeal.representative or ""],
                    ["Section", appeal.section or ""],
                    ["Status", appeal.status or ""],
                    ["Tax Date", appeal.tax_date or ""],
                    ["Hearing Number", appeal.hearing_number or ""],
                    ["Hearing Date", appeal.hearing_date or ""],
                    ["Board Order Number", appeal.board_order_number or ""],
                    ["Filing Date", appeal.filing_date or ""],
                    ["Reason for Appeal", appeal.reason_for_appeal or ""],
                    ["Decision Number", appeal.decision_number or ""],
                    ["Decision Mailing Date", appeal.decision_mailing_date or ""],
                    ["Decisions", appeal.decisions or ""],
                    ["Decision Details", appeal.decision_details or ""],
                    ["Property Roll Number", appeal.property_roll_number or ""],
                    ["Property Municipality", appeal.property_municipality or ""],
                    ["Property Classification", appeal.property_classification or ""],
                    ["Property Neighborhood", appeal.property_nbhd or ""],
                    ["Property Description", appeal.property_description or ""],
                    ["Created", appeal.created_at.strftime("%Y-%m-%d %H:%M:%S") if appeal.created_at else ""],
                    ["Updated", appeal.updated_at.strftime("%Y-%m-%d %H:%M:%S") if appeal.updated_at else ""]
                ]
                
                for row_idx, (field, value) in enumerate(appeal_data, 2):
                    appeal_sheet.cell(row=row_idx, column=1, value=field)
                    appeal_sheet.cell(row=row_idx, column=2, value=str(value))
                
                # Adjust column widths for appeal sheet
                appeal_sheet.column_dimensions["A"].width = 30
                appeal_sheet.column_dimensions["B"].width = 50
            
            # Save Excel file
            safe_filename = roll_record.roll_number.replace('/', '_').replace('\\', '_').replace(':', '_')
            excel_path = temp_path / f"{safe_filename}.xlsx"
            wb.save(excel_path)
        
        # Create ZIP file
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for excel_file in temp_path.glob("*.xlsx"):
                zip_file.write(excel_file, excel_file.name)
        
        zip_buffer.seek(0)
        
        await add_log("INFO", f"Exported {len(roll_numbers)} roll numbers with appeals as ZIP")
        
        return StreamingResponse(
            zip_buffer,
            media_type="application/zip",
            headers={
                "Content-Disposition": f"attachment; filename=roll_numbers_with_appeals_{type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
            }
        )


# Server-Sent Events for real-time updates
@app.get("/api/sse/logs")
async def sse_logs(request: Request, _: bool = Depends(require_auth)):
    """SSE endpoint for real-time logs."""
    async def event_generator():
        queue = asyncio.Queue()
        log_subscribers.append(queue)
        
        try:
            while True:
                # Check if client disconnected
                if await request.is_disconnected():
                    break
                
                # Wait for new log
                try:
                    log_entry = await asyncio.wait_for(queue.get(), timeout=30)
                    yield f"data: {json.dumps(log_entry)}\n\n"
                except asyncio.TimeoutError:
                    # Send heartbeat
                    yield f"data: {json.dumps({'type': 'heartbeat'})}\n\n"
                    
        finally:
            log_subscribers.remove(queue)
    
    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
        }
    )


@app.get("/api/sse/extraction-progress")
async def sse_extraction_progress(request: Request, _: bool = Depends(require_auth)):
    """SSE endpoint for extraction progress."""
    async def event_generator():
        while True:
            if await request.is_disconnected():
                break
                
            # Send current extraction status
            yield f"data: {json.dumps(active_extractions)}\n\n"
            
            await asyncio.sleep(1)  # Update every second
    
    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
        }
    )


# Database management endpoints
@app.get("/api/database/info")
async def get_database_info(db: Session = Depends(get_db), _: bool = Depends(require_auth)):
    """Get database information including size."""
    import os
    from src.creiq.database.database import DATABASE_URL
    
    db_info = {
        "size": "Unknown",
        "type": "PostgreSQL" if DATABASE_URL.startswith("postgresql") else "SQLite",
        "location": DATABASE_URL.split("///")[-1] if "///" in DATABASE_URL else "Remote",
        "backups": []
    }
    
    # Get database size
    if DATABASE_URL.startswith("sqlite"):
        db_path = DATABASE_URL.split("///")[-1]
        if os.path.exists(db_path):
            size_bytes = os.path.getsize(db_path)
            # Convert to human readable
            for unit in ['B', 'KB', 'MB', 'GB']:
                if size_bytes < 1024.0:
                    db_info["size"] = f"{size_bytes:.1f} {unit}"
                    break
                size_bytes /= 1024.0
    else:
        # For PostgreSQL, we'd need to run a query
        try:
            result = db.execute("SELECT pg_database_size(current_database())")
            size_bytes = result.scalar()
            for unit in ['B', 'KB', 'MB', 'GB']:
                if size_bytes < 1024.0:
                    db_info["size"] = f"{size_bytes:.1f} {unit}"
                    break
                size_bytes /= 1024.0
        except:
            pass
    
    # Get list of backups
    backup_dir = Path("backups")
    if backup_dir.exists():
        backups = sorted(
            [f for f in backup_dir.glob("*.db") or backup_dir.glob("*.sql")],
            key=lambda x: x.stat().st_mtime,
            reverse=True
        )[:3]  # Last 3 backups
        
        db_info["backups"] = [
            {
                "filename": f.name,
                "size": f"{f.stat().st_size / 1024 / 1024:.1f} MB",
                "created": datetime.fromtimestamp(f.stat().st_mtime).isoformat()
            }
            for f in backups
        ]
    
    return db_info


@app.post("/api/database/backup")
async def backup_database(db: Session = Depends(get_db), _: bool = Depends(require_auth)):
    """Create a backup of the database."""
    from src.creiq.database.database import DATABASE_URL
    import shutil
    
    backup_dir = Path("backups")
    backup_dir.mkdir(exist_ok=True)
    
    # First, check if we need to delete old backups
    existing_backups = sorted(
        [f for f in backup_dir.glob("*.db") if f.is_file()],
        key=lambda x: x.stat().st_mtime,
        reverse=False  # Oldest first
    )
    
    # If we already have 3 or more backups, delete the oldest one
    if len(existing_backups) >= 3:
        # Delete the oldest backup (first in the list since reverse=False)
        oldest_backup = existing_backups[0]
        try:
            oldest_backup.unlink()
            await add_log("INFO", f"Removed oldest backup: {oldest_backup.name}")
        except Exception as e:
            await add_log("WARNING", f"Failed to remove old backup {oldest_backup.name}: {str(e)}")
    
    # Now create the new backup
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    try:
        if DATABASE_URL.startswith("sqlite"):
            # SQLite backup - simple file copy
            db_path = DATABASE_URL.split("///")[-1]
            backup_path = backup_dir / f"creiq_backup_{timestamp}.db"
            
            # Close all connections before copying
            db.close()
            engine.dispose()
            
            shutil.copy2(db_path, backup_path)
            
            await add_log("INFO", f"Database backup created: {backup_path.name}")
            
            return {
                "success": True,
                "filename": backup_path.name,
                "size": f"{backup_path.stat().st_size / 1024 / 1024:.1f} MB"
            }
            
        else:
            # PostgreSQL backup would require pg_dump
            raise HTTPException(status_code=501, detail="PostgreSQL backup not implemented yet")
        
    except Exception as e:
        await add_log("ERROR", f"Failed to create backup: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/database/backup/{filename}")
async def download_backup(filename: str, _: bool = Depends(require_auth)):
    """Download a database backup."""
    backup_path = Path("backups") / filename
    
    if not backup_path.exists() or not backup_path.is_file():
        raise HTTPException(status_code=404, detail="Backup not found")
    
    return StreamingResponse(
        open(backup_path, "rb"),
        media_type="application/octet-stream",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )


@app.delete("/api/database/backup/{filename}")
async def delete_backup(filename: str, _: bool = Depends(require_auth)):
    """Delete a specific database backup."""
    backup_path = Path("backups") / filename
    
    # Security check - ensure the file is in the backups directory
    if not backup_path.exists() or not backup_path.is_file():
        raise HTTPException(status_code=404, detail="Backup not found")
    
    # Ensure the path is within the backups directory (prevent directory traversal)
    try:
        backup_path = backup_path.resolve()
        backup_dir = Path("backups").resolve()
        if not str(backup_path).startswith(str(backup_dir)):
            raise HTTPException(status_code=403, detail="Invalid backup path")
    except Exception:
        raise HTTPException(status_code=403, detail="Invalid backup path")
    
    try:
        backup_path.unlink()
        await add_log("INFO", f"Deleted backup: {filename}")
        return {"success": True, "message": f"Backup {filename} deleted successfully"}
    except Exception as e:
        await add_log("ERROR", f"Failed to delete backup {filename}: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/database/purge")
async def purge_database(
    request: Request,
    passcode: str = Form(...),
    db: Session = Depends(get_db),
    _: bool = Depends(require_auth)
):
    """Purge all data from the database (requires passcode confirmation)."""
    # Verify passcode
    if passcode != PASSCODE:
        raise HTTPException(status_code=403, detail="Invalid passcode")
    
    try:
        # Delete all data in reverse order of dependencies
        db.query(Appeal).delete()
        db.query(RollNumber).delete()
        db.commit()
        
        # Clear active extractions
        active_extractions.clear()
        
        await add_log("WARNING", "Database purged - all data deleted")
        
        return {"success": True, "message": "Database purged successfully"}
        
    except Exception as e:
        db.rollback()
        await add_log("ERROR", f"Failed to purge database: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/database/import")
async def import_database(
    passcode: str = Form(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _: bool = Depends(require_auth)
):
    """Import and replace the database (requires passcode confirmation)."""
    from src.creiq.database.database import DATABASE_URL
    import shutil
    
    # Verify passcode
    if passcode != PASSCODE:
        raise HTTPException(status_code=403, detail="Invalid passcode")
    
    if not file.filename.endswith('.db'):
        raise HTTPException(status_code=400, detail="Only .db files are allowed")
    
    try:
        if DATABASE_URL.startswith("sqlite"):
            # Create a backup first
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_dir = Path("backups")
            backup_dir.mkdir(exist_ok=True)
            
            db_path = DATABASE_URL.split("///")[-1]
            backup_path = backup_dir / f"creiq_before_import_{timestamp}.db"
            
            # Close all connections
            db.close()
            engine.dispose()
            
            # Backup current database
            shutil.copy2(db_path, backup_path)
            
            # Save uploaded file to temp location
            temp_path = Path(f"temp_import_{timestamp}.db")
            with open(temp_path, "wb") as buffer:
                shutil.copyfileobj(file.file, buffer)
            
            # Replace database
            shutil.move(str(temp_path), db_path)
            
            await add_log("WARNING", f"Database replaced with imported file: {file.filename}")
            
            return {
                "success": True,
                "message": "Database imported successfully",
                "backup_created": backup_path.name
            }
            
        else:
            raise HTTPException(status_code=501, detail="PostgreSQL import not implemented yet")
            
    except Exception as e:
        await add_log("ERROR", f"Failed to import database: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/roll_numbers/{roll_number}")
async def delete_roll_number(
    roll_number: str,
    db: Session = Depends(get_db),
    _: bool = Depends(require_auth)
):
    """
    Delete a roll number and all associated appeals.
    
    Args:
        roll_number: The roll number to delete
        
    Returns:
        Deletion confirmation
    """
    db_service = DatabaseService(db)
    
    if not db_service.delete_roll_number(roll_number):
        raise HTTPException(status_code=404, detail="Roll number not found")
    
    await add_log("INFO", f"Deleted roll number: {roll_number}")
    return {"message": f"Roll number {roll_number} and associated appeals deleted successfully"}


@app.delete("/api/appeals/{appeal_id}")
async def delete_appeal(
    appeal_id: int,
    db: Session = Depends(get_db),
    _: bool = Depends(require_auth)
):
    """
    Delete an individual appeal.
    
    Args:
        appeal_id: The appeal ID to delete
        
    Returns:
        Deletion confirmation
    """
    db_service = DatabaseService(db)
    
    if not db_service.delete_appeal(appeal_id):
        raise HTTPException(status_code=404, detail="Appeal not found")
    
    await add_log("INFO", f"Deleted appeal: {appeal_id}")
    return {"message": f"Appeal {appeal_id} deleted successfully"}


if __name__ == "__main__":
    # Ensure directories exist
    templates_path.mkdir(exist_ok=True)
    static_path.mkdir(exist_ok=True)
    (static_path / "css").mkdir(exist_ok=True)
    (static_path / "js").mkdir(exist_ok=True)
    (static_path / "img").mkdir(exist_ok=True)
    
    # Run the server
    uvicorn.run(
        "src.creiq.web_app:app",
        host=API_HOST,
        port=API_PORT,
        reload=API_RELOAD
    ) 